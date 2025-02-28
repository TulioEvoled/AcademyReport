from flask import Blueprint, request, jsonify, send_file, render_template, redirect, url_for
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
from io import BytesIO, StringIO
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font
import win32com.client
import os
import pythoncom
from datetime import datetime
from routes.auth_routes import login_required

# Definir el Blueprint para Ingeniería en Informática
informatica_bp = Blueprint('informatica', __name__, url_prefix='/informatica')

# Conexión a MongoDB
client = MongoClient('mongodb+srv://ivan:tuliogaymer077@cluster0.bkahq7u.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0')
db = client.tecnologico

# Colecciones específicas de Ingeniería en Informática
informatica_profesores = db['informatica_profesores']
informatica_asignaturas = db['informatica_asignaturas']
informatica_asignaturasE = db['informatica_asignaturasE']
administrativos = db['administrativos']

# Listas de grupos, horarios y carreras
informatica_grupos = [
    "NG", "6101", "6102", "6151", "6152", "6201", "6202", "6251", "6252",
    "6301", "6302", "6351", "6352", "6401", "6402", "6451", "6452", "6501", 
    "6502", "6551", "6552", "6601", "6602", "6651", "6652", "6751", "6752", 
    "6851", "6852", "6951", "6952"
]

horarios = [
    "07:00", "08:00", "09:00", "10:00", "11:00", "12:00", "13:00",
    "14:00", "15:00", "16:00", "17:00", "18:00", "19:00", "20:00",
    "21:00"
]

carreras = [
    "INFORMÁTICA", "INDUSTRIAL", "ELECTRÓNICA",
    "ELECTROMECÁNICA", "SISTEMAS COMPUTACIONALES", "ADMINISTRACIÓN"
]

# Configuración de carga de archivos para la carrera de Ingeniería en Informática
INFORMATICA_UPLOAD_FOLDER = "informatica/static/src/"
INFORMATICA_ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

# Rutas CRUD para profesores en Ingeniería en Informática
@informatica_bp.route('/informatica_profesores', methods=['POST'])
@login_required('Informatica')
def add_profesor():
    data = request.json
    informatica_profesores.insert_one(data)
    return jsonify({'msg': 'Profesor añadido'}), 201

@informatica_bp.route('/informatica_profesores/Add_Profesor')
@login_required('Informatica')
def add_profesor2():
    return render_template('informatica/Add_Profesor.html')

@informatica_bp.route('/informatica_profesores/<id>', methods=['GET'])
@login_required('Informatica')
def get_profesor(id):
    profesor = informatica_profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return jsonify(profesor)

@informatica_bp.route('/informatica_profesores/<id>', methods=['PUT'])
@login_required('Informatica')
def update_profesor(id):
    data = request.json
    informatica_profesores.update_one({'_id': ObjectId(id)}, {'$set': data})
    return jsonify({'msg': 'Profesor actualizado'})

@informatica_bp.route('/informatica_profesores/edit/<id>', methods=['GET'])
@login_required('Informatica')
def edit_profesor(id):
    profesor = informatica_profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return render_template('informatica/edit_profesor.html', profesor=profesor, grupos=informatica_grupos, horarios=horarios)

@informatica_bp.route('/informatica_profesores/<id>', methods=['DELETE'])
@login_required('Informatica')
def delete_profesor(id):
    informatica_profesores.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Profesor eliminado'})

@informatica_bp.route('/informatica_profesores', methods=['GET'])
@login_required('Informatica')
def get_all_profesores():
    all_profesores = list(informatica_profesores.find({}).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return render_template('informatica/profesores.html', profesores=all_profesores)

# Ruta para obtener la lista de profesores en formato JSON en Ingeniería en Informática
@informatica_bp.route('/informatica_profesores/json', methods=['GET'])
@login_required('Informatica')
def get_profesores_json():
    all_profesores = list(informatica_profesores.find({}, {'_id': 1, 'nombre': 1}).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return jsonify(all_profesores)

# Rutas CRUD para asignaturas en Ingeniería en Informática
@informatica_bp.route('/informatica_asignaturas', methods=['POST'])
@login_required('Informatica')
def add_asignatura():
    data = request.json
    informatica_asignaturas.insert_one(data)
    return jsonify({'msg': 'Asignatura añadida'}), 201

@informatica_bp.route('/informatica_asignaturas/<id>', methods=['GET'])
@login_required('Informatica')
def get_asignatura(id):
    asignatura = informatica_asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return jsonify(asignatura)

@informatica_bp.route('/informatica_asignaturas/<id>', methods=['PUT'])
@login_required('Informatica')
def update_asignatura(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = informatica_asignaturas.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Asignatura actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar la asignatura"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@informatica_bp.route('/informatica_edit-asignatura/<id>', methods=['GET'])
@login_required('Informatica')
def edit_asignatura(id):
    asignatura = informatica_asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return render_template('informatica/edit_asignatura.html', asignatura=asignatura)

@informatica_bp.route('/informatica_asignaturas/<id>', methods=['DELETE'])
@login_required('Informatica')
def delete_asignatura(id):
    informatica_asignaturas.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Asignatura eliminada'})

@informatica_bp.route('/informatica_asignaturas/json', methods=['GET'])
@login_required('Informatica')
def get_all_asignaturas_json():
    all_asignaturas = list(informatica_asignaturas.find({}))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return jsonify(all_asignaturas)

@informatica_bp.route('/informatica_asignaturas', methods=['GET'])
@login_required('Informatica')
def get_all_asignaturas():
    all_asignaturas = list(informatica_asignaturas.find({}).sort("nombre", 1))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return render_template('informatica/asignaturas.html', asignaturas=all_asignaturas)

# Rutas CRUD para asignaturas especiales en Ingeniería en Informática
@informatica_bp.route('/informatica_asignaturasE', methods=['POST'])
@login_required('Informatica')
def add_asignaturaE():
    data = request.json
    informatica_asignaturasE.insert_one(data)
    return jsonify({'msg': 'Asignatura Especial añadida'}), 201

@informatica_bp.route('/informatica_asignaturasE/<id>', methods=['GET'])
@login_required('Informatica')
def get_asignaturaE(id):
    asignaturaE = informatica_asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(asignaturaE)

@informatica_bp.route('/informatica_asignaturasE/<id>', methods=['PUT'])
@login_required('Informatica')
def update_asignaturaE(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = informatica_asignaturasE.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Asignatura Especial actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar la Asignatura Especial"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@informatica_bp.route('/informatica_edit-asignaturaE/<id>', methods=['GET'])
@login_required('Informatica')
def edit_asignaturaE(id):
    asignaturaE = informatica_asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('informatica/edit_asignaturaE.html', asignaturaE=asignaturaE)

@informatica_bp.route('/informatica_asignaturasE/<id>', methods=['DELETE'])
@login_required('Informatica')
def delete_asignaturaE(id):
    informatica_asignaturasE.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Asignatura Especial eliminada'})

@informatica_bp.route('/informatica_asignaturasE/json', methods=['GET'])
@login_required('Informatica')
def get_all_asignaturasE_json():
    all_asignaturasE = list(informatica_asignaturasE.find({}))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(all_asignaturasE)

@informatica_bp.route('/informatica_asignaturasE', methods=['GET'])
@login_required('Informatica')
def get_all_asignaturasE():
    all_asignaturasE = list(informatica_asignaturasE.find({}).sort("nombre", 1))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('informatica/asignaturasE.html', asignaturasE=all_asignaturasE)

# Función para exportar datos en Ingeniería en Informática
# Rutas para obtener las columnas disponibles
@informatica_bp.route('/informatica_columns/<collection_name>', methods=['GET'])
@login_required('Informatica')
def get_columns(collection_name):
    column_mappings = {
        "profesores": [
            "nombre", "profesion", "adscripcion", "fecha_ingreso",
            "tiempo_indeterminado", "periodo_actual", "horas_a",
            "horas_b", "Horas de Asignatura", "Horas Descarga", "total_horas"
        ],
        "asignaturas": ["nombre", "horas"],
        "asignaturasE": ["nombre", "horas"]  # Nueva colección de Asignaturas Especiales
    }

    if collection_name not in column_mappings:
        return jsonify({"error": "Colección no encontrada"}), 404

    return jsonify(column_mappings[collection_name])

# Función para exportar datos
@informatica_bp.route('/informatica_export', methods=['POST'])
@login_required('Informatica')
def export_data():
    data = request.json
    selected_columns = data['columns'].split(',')
    collection_name = data['collection']
    export_format = data.get('format', 'xlsx')
    
     # Seleccionar la colección correcta
    if collection_name == "profesores":
        collection = informatica_profesores
    elif collection_name == "asignaturas":
        collection = informatica_asignaturas
    elif collection_name == "asignaturasE":
        collection = informatica_asignaturasE
    else:
        return jsonify({"error": "Colección no válida"}), 400

    cursor = collection.find({})
    df = pd.DataFrame(list(cursor))

    # Reemplazar valores NaN e Inf en el DataFrame antes de exportarlo
    df = df.replace([float('inf'), -float('inf')], 0).fillna('')

    if df.empty:
        return {"message": "No hay datos para exportar"}, 400

    # Ordenar alfabéticamente por "nombre"
    df = df.sort_values(by="nombre", ascending=True)

    # **📌 Si el usuario selecciona estas columnas, las calculamos**
    if "Horas de Asignatura" in selected_columns:
        df["Horas de Asignatura"] = df.apply(lambda row: sum(
            int(row.get(f"horas{i}", 0) or 0) for i in range(1, 9) if row.get(f"carrera{i}") == "INFORMÁTICA"
        ), axis=1)

    if "Horas Descarga" in selected_columns:
        df["Horas Descarga"] = df.apply(lambda row: sum(
            int(row.get(f"horasE{i}", 0) or 0) for i in range(1, 9) if row.get(f"carreraE{i}") == "INFORMÁTICA"
        ), axis=1)

    output = BytesIO()

    if export_format == 'xlsx':
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            sheet = workbook.add_worksheet('Datos')

            # Definir colores alternos
            format_white = workbook.add_format({'bg_color': '#FFFFFF'})  # Blanco
            format_gray = workbook.add_format({'bg_color': '#F2F2F2'})  # Gris claro
            format_header = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF'})
            format_red = workbook.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})  # Rojo para errores

            # **📌 Filtrar solo las columnas seleccionadas**
            selected_columns_filtered = [col for col in selected_columns if col not in [
                "asignacion_horas_frente_grupo", "asignacion_horas_descarga_otras_actividades", "asignacion_horas_cargo_academico"
            ]]

            # Escribir encabezados
            sheet.write_row(0, 0, selected_columns_filtered, format_header)

            # Escribir datos con colores alternos
            for row_num, row in enumerate(df[selected_columns_filtered].values, start=1):
                color_format = format_gray if row_num % 2 == 0 else format_white

                # Escribir toda la fila normalmente
                sheet.write_row(row_num, 0, row, color_format)

                # Verificar si "total_horas" debe ser pintado de rojo
                if "total_horas" in selected_columns_filtered:
                    total_horas_index = selected_columns_filtered.index("total_horas")
                    total_horas = int(row[total_horas_index]) if row[total_horas_index] else 0
                    horas_asignatura = int(row[selected_columns_filtered.index("Horas de Asignatura")]) if "Horas de Asignatura" in selected_columns_filtered else 0
                    horas_descarga = int(row[selected_columns_filtered.index("Horas Descarga")]) if "Horas Descarga" in selected_columns_filtered else 0

                    # Verificar si la diferencia es distinta de 0
                    if (horas_asignatura + horas_descarga) != total_horas:
                        sheet.write(row_num, selected_columns_filtered.index("total_horas"), total_horas, format_red)

            # **📌 1. Horas Frente a Grupo**
            if 'asignacion_horas_frente_grupo' in selected_columns:
                sheet_horarios = workbook.add_worksheet("Horas Frente a Grupo")
                headers = ["Carrera", "Asignatura", "Grupo", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                sheet_horarios.write_row(0, 0, headers, format_header)

                row_num = 1
                for index, profesor in enumerate(df.to_dict(orient="records")):
                    color_format = format_gray if index % 2 == 0 else format_white
                    for i in range(1, 9):
                        sheet_horarios.write(row_num, 0, profesor.get(f"carrera{i}", ""), color_format)
                        sheet_horarios.write(row_num, 1, profesor.get(f"asignatura{i}", ""), color_format)
                        sheet_horarios.write(row_num, 2, profesor.get(f"grupo{i}", ""), color_format)
                        sheet_horarios.write(row_num, 3, profesor.get(f"horas{i}", 0), color_format)

                        # Horarios por día (Lunes - Sábado)
                        for j in range(1, 7):  
                            horario_inicio = profesor.get(f"hora_inicio{i}{j}", "")
                            horario_fin = profesor.get(f"hora_fin{i}{j}", "")
                            sheet_horarios.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                        row_num += 1

            # **📌 2. Horas Descarga Otras Actividades**
            if 'asignacion_horas_descarga_otras_actividades' in selected_columns:
                sheet_descarga = workbook.add_worksheet("Horas Descarga")
                headers_descarga = ["Carrera", "Asignatura", "Grupo", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                sheet_descarga.write_row(0, 0, headers_descarga, format_header)

                row_num = 1
                for index, profesor in enumerate(df.to_dict(orient="records")):
                    color_format = format_gray if index % 2 == 0 else format_white
                    for i in range(1, 9):
                        sheet_descarga.write(row_num, 0, profesor.get(f"carreraE{i}", ""), color_format)
                        sheet_descarga.write(row_num, 1, profesor.get(f"asignaturaE{i}", ""), color_format)
                        sheet_descarga.write(row_num, 2, profesor.get(f"grupoE{i}", ""), color_format)
                        sheet_descarga.write(row_num, 3, profesor.get(f"horasE{i}", 0), color_format)

                        for j in range(1, 7):
                            horario_inicio = profesor.get(f"hora_inicioE{i}{j}", "")
                            horario_fin = profesor.get(f"hora_finE{i}{j}", "")
                            sheet_descarga.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                        row_num += 1

            # **📌 3. Horas Cargo Académico**
            if 'asignacion_horas_cargo_academico' in selected_columns:
                sheet_cargo = workbook.add_worksheet("Horas Cargo Académico")
                headers_cargo = ["Carrera", "Cargo", "Vigencia", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                sheet_cargo.write_row(0, 0, headers_cargo, format_header)

                row_num = 1
                for index, profesor in enumerate(df.to_dict(orient="records")):
                    color_format = format_gray if index % 2 == 0 else format_white
                    sheet_cargo.write(row_num, 0, profesor.get("carreraC", ""), color_format)
                    sheet_cargo.write(row_num, 1, profesor.get("cargo", ""), color_format)
                    sheet_cargo.write(row_num, 2, profesor.get("vigenciaCargo", ""), color_format)
                    sheet_cargo.write(row_num, 3, profesor.get("horasC", 0), color_format)

                    for j in range(1, 7):
                        horario_inicio = profesor.get(f"hora_inicioC1{j}", "")
                        horario_fin = profesor.get(f"hora_finC1{j}", "")
                        sheet_cargo.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                    row_num += 1

        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"{collection_name}.xlsx")

# Rutas de la interfaz de usuario para Ingeniería en Informática
@informatica_bp.route('/informatica_index')
@login_required('Informatica')
def index():
    return render_template('informatica/index.html')

@informatica_bp.route('/informatica')
@login_required('Informatica')
def principal():
    return render_template('informatica/principal.html')

@informatica_bp.route('/informatica_reporteador')
@login_required('Informatica')
def reporteador():
    return render_template('informatica/exportacion/exportar.html')

# Validar archivos permitidos para imágenes
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# Ruta para subir imágenes de encabezado y pie de página en Informática
@informatica_bp.route("/informatica_upload-images", methods=["POST"])
@login_required('Informatica')
def upload_images():
    upload_folder = "static/informatica/src/"

    if "header" in request.files:
        header = request.files["header"]
        if header and allowed_file(header.filename):
            header.save(os.path.join(upload_folder, "Encabezado1.PNG"))

    if "footer" in request.files:
        footer = request.files["footer"]
        if footer and allowed_file(footer.filename):
            footer.save(os.path.join(upload_folder, "PieDePagina1.PNG"))

    return jsonify({'msg': 'Imágenes actualizadas correctamente'}), 200

# Ruta para actualizar el texto en la celda A4 en Informática
@informatica_bp.route('/informatica_update-text', methods=['POST'])
@login_required('Informatica')
def update_text():
    nuevo_texto = request.form.get('nuevo_texto', '')
    if not nuevo_texto:
        return jsonify({'msg': 'Error: No se proporcionó un texto válido'}), 400
    
    # Guardar el nuevo texto en un archivo de configuración
    texto_a4_path = "static/informatica/src/texto_a4.txt"
    with open(texto_a4_path, "w", encoding="utf-8") as file:
        file.write(nuevo_texto)
    
    return jsonify({'msg': 'Texto A4 actualizado correctamente'}), 200

# Ruta para actualizar el texto en la celda A54 en Informática
@informatica_bp.route('/informatica_update-text-a54', methods=['POST'])
@login_required('Informatica')
def update_text_a54():
    nuevo_texto_dos = request.form.get('nuevo_texto_dos', '')
    if not nuevo_texto_dos:
        return jsonify({'msg': 'Error: No se proporcionó un texto válido'}), 400
    
    # Guardar el nuevo texto en un archivo de configuración
    texto_a54_path = "static/informatica/src/texto_a54.txt"
    with open(texto_a54_path, "w", encoding="utf-8") as file:
        file.write(nuevo_texto_dos)
    
    return jsonify({'msg': 'Texto A54 actualizado correctamente'}), 200

# Ruta para exportar los profesores seleccionados usando la plantilla con 32 hojas EXCEL en Informática
@informatica_bp.route('/informatica_export-selected', methods=['POST'])
@login_required('Informatica')
def export_selected():
    profesor_ids = request.form.getlist('profesor_ids')
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [informatica_profesores.find_one({'_id': ObjectId(profesor_id)}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas
    template_path = "static/informatica/src/Plantilla_pie_reducido_2cm.xlsx"
    workbook = openpyxl.load_workbook(template_path)

    # Cargar imágenes
    header_image = Image("static/informatica/src/Encabezado1.PNG")
    footer_image = Image("static/informatica/src/PieDePagina1.PNG")

    # Establecer bordes
    thin_side = Side(style='thin')
    border_A44 = Border(bottom=thin_side, left=thin_side)
    border_B44 = Border(bottom=thin_side)

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 y A54 desde archivos de configuración
    texto_a4_path = "static/informatica/src/texto_a4.txt"
    texto_a54_path = "static/informatica/src/texto_a54.txt"
    
    texto_a4 = open(texto_a4_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a4_path) else ""
    texto_a54 = open(texto_a54_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a54_path) else ""

    # Obtener el nombre y cargo del encargado de dirección académica y general
    encargado = administrativos.find_one({'cargo': {'$in': ["ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA"]}})
    direccion = administrativos.find_one({'cargo': {'$in': ["DIRECTORA GENERAL", "DIRECTOR GENERAL"]}})

    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos en Informática
    cargo_mapping = {
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES"],
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "ELECTROMECÁNICA": ["JEFA DE DIVISIÓN DE ING. ELECTROMECÁNICA", "JEFE DE DIVISIÓN DE ING. ELECTROMECÁNICA"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        # Insertar imágenes en encabezado y pie de página
        sheet.add_image(header_image, "A1")
        sheet.add_image(footer_image, "A56")

        # Actualizar el texto en la celda A4 y A54
        sheet["A4"] = texto_a4
        sheet["A54"] = texto_a54

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de dirección
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("2"):
                    sheet[f"A{row}"] = "ELECTROMECÁNICA"
                elif str(grupo).startswith("3"):
                    sheet[f"A{row}"] = "ELECTRÓNICA"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"
                elif str(grupo).startswith("6"):
                    sheet[f"A{row}"] = "INFORMÁTICA"
                elif str(grupo).startswith("9"):
                    sheet[f"A{row}"] = "ADMINISTRACIÓN"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)

        # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado = encargado["nombre"] if encargado else ""
            cargo_encargado = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado
            sheet["A46"] = nombre_encargado
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_A44
            sheet["B44"].border = border_B44

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # Guardar el archivo en memoria
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Devolver el archivo actualizado como respuesta
    return send_file(
        output,
        as_attachment=True,
        download_name="Reporte_Profesores_Informatica.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Función para convertir Excel a PDF usando win32com en Informática
def excel_to_pdf_informatica(input_excel_path, output_pdf_path):
    pythoncom.CoInitialize()  # Inicializa el entorno COM
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Ejecutar Excel en segundo plano

        workbook = excel.Workbooks.Open(os.path.abspath(input_excel_path))
        
        # Iterar por todas las hojas y deshabilitar encabezados/pies de página
        for sheet in workbook.Sheets:
            sheet.PageSetup.CenterFooter = ""  # Elimina el pie de página central
            sheet.PageSetup.LeftFooter = ""    # Elimina el pie de página izquierdo
            sheet.PageSetup.RightFooter = ""   # Elimina el pie de página derecho
            sheet.PageSetup.CenterHeader = ""  # Elimina el encabezado central
            sheet.PageSetup.LeftHeader = ""    # Elimina el encabezado izquierdo
            sheet.PageSetup.RightHeader = ""   # Elimina el encabezado derecho
        
        # Exportar como PDF sin encabezado ni pie de página
        workbook.ExportAsFixedFormat(0, os.path.abspath(output_pdf_path))
        workbook.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()  # Liberar el entorno COM

# Ruta para exportar los profesores seleccionados en PDF en Informática
@informatica_bp.route('/informatica_export-selected-pdf', methods=['POST'])
@login_required('Informatica')
def export_selected_pdf_informatica():
    profesor_ids = request.form.getlist('profesor_ids')

    if not profesor_ids:
        return "Error: No se recibieron IDs de profesores", 400
    
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [informatica_profesores.find_one({'_id': ObjectId(profesor_id)}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas en Informática
    template_path = "static/informatica/src/Plantilla_pie_reducido_2cm.xlsx"
    workbook = openpyxl.load_workbook(template_path)

    # Cargar imágenes
    header_image = Image("static/informatica/src/Encabezado1.PNG")
    footer_image = Image("static/informatica/src/PieDePagina1.PNG")

    # Establecer bordes
    thin_side = Side(style='thin')
    border_A44 = Border(bottom=thin_side, left=thin_side)
    border_B44 = Border(bottom=thin_side)

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 y A54 desde archivos de configuración
    texto_a4_path = "static/informatica/src/texto_a4.txt"
    texto_a54_path = "static/informatica/src/texto_a54.txt"
    
    texto_a4 = open(texto_a4_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a4_path) else ""
    texto_a54 = open(texto_a54_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a54_path) else ""

    # Obtener el nombre y cargo del encargado de dirección académica y general
    encargado = administrativos.find_one({'cargo': {'$in': ["ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA"]}})
    direccion = administrativos.find_one({'cargo': {'$in': ["DIRECTORA GENERAL", "DIRECTOR GENERAL"]}})

    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos en Informática
    cargo_mapping = {
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES"],
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "ELECTROMECÁNICA": ["JEFA DE DIVISIÓN DE ING. ELECTROMECÁNICA", "JEFE DE DIVISIÓN DE ING. ELECTROMECÁNICA"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        # Insertar imágenes en encabezado y pie de página
        sheet.add_image(header_image, "A1")
        sheet.add_image(footer_image, "A56")

        # Actualizar el texto en la celda A4 y A54
        sheet["A4"] = texto_a4
        sheet["A54"] = texto_a54

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de dirección
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("2"):
                    sheet[f"A{row}"] = "ELECTROMECÁNICA"
                elif str(grupo).startswith("3"):
                    sheet[f"A{row}"] = "ELECTRÓNICA"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"
                elif str(grupo).startswith("6"):
                    sheet[f"A{row}"] = "INFORMÁTICA"
                elif str(grupo).startswith("9"):
                    sheet[f"A{row}"] = "ADMINISTRACIÓN"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)

        # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado = encargado["nombre"] if encargado else ""
            cargo_encargado = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado
            sheet["A46"] = nombre_encargado
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_A44
            sheet["B44"].border = border_B44

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # Guardar el archivo Excel temporalmente para la conversión a PDF
    temp_excel_path = "temp_reporte_informatica.xlsx"
    workbook.save(temp_excel_path)
    workbook.close()

    # Convertir el archivo Excel a PDF
    pdf_path = "Reporte_Profesores_Informatica.pdf"
    excel_to_pdf_informatica(temp_excel_path, pdf_path)

    # Enviar el archivo PDF como respuesta
    return send_file(
        pdf_path,
        as_attachment=True,
        download_name="Reporte_Profesores_Informatica.pdf",
        mimetype="application/pdf"
    )