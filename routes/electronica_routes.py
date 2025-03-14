from flask import Blueprint, request, jsonify, send_file, render_template, redirect, url_for, session, send_from_directory, Response
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
from io import BytesIO, StringIO
import io
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font
import win32com.client
import os
import pythoncom
from datetime import datetime
import subprocess
import time
import gridfs
from routes.auth_routes import login_required

# Definir el Blueprint para Ingeniería en Electrónica
electronica_bp = Blueprint('electronica', __name__, url_prefix='/electronica')

# Conexión a MongoDB
client = MongoClient('mongodb+srv://ivan:tuliogaymer077@cluster0.bkahq7u.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0')
db = client.tecnologico
fs = gridfs.GridFS(db)  # Inicializa GridFS para almacenamiento de archivos

# Colecciones específicas de Ingeniería en Electrónica
electronica_profesores = db['profesores']
electronica_asignaturas = db['electronica_asignaturas']
electronica_asignaturasE = db['electronica_asignaturasE']
administrativos = db['administrativos']

# Listas de grupos, horarios y carreras
electronica_grupos = [
    #INDUSTRIAL
    "NG", "1101", "1102", "1151", "1152", "1181", "1201", "1202", "1251", "1252", "1281", "1301", "1302", "1351", "1352", "1381", "1401", "1402", "1451",
    "1452", "1481", "1501", "1502", "1551", "1552", "1581", "1601", "1602", "1651", "1652", "1681", "1751", "1752", "1781", "1851", "1852", "1881",
    "1951", "1952", "1981",
    #SISTEMAS
    "4101", "4102", "4151", "4152", "4171", "4201", "4202", "4251", "4252", "4271", "4301", "4302", "4351", "4352", "4371", "4401", "4402", "4451",
    "4452", "4471", "4501", "4502", "4551", "4552", "4571", "4601", "4602", "4651", "4652", "4671", "4751", "4752", "4771", "4851", "4852", "4871",
    "4951", "4952", "4971",
    #INFORMATICA
    "6101", "6102", "6151", "6152", "6201", "6202", "6251", "6252", "6301", "6302", "6351", "6352", "6401", "6402", "6451", "6452", "6501", 
    "6502", "6551", "6552", "6601", "6602", "6651", "6652", "6751", "6752", "6851", "6852", "6951", "6952",
    #ELECTRONICA
    "3101", "3102", "3151", "3152", "3201", "3202", "3251", "3252", "3301", "3302", "3351", "3352", "3401", "3402", "3451", "3452", "3501", 
    "3502", "3551", "3552", "3601", "3602", "3651", "3652", "3751", "3752", "3851", "3852", "3951", "3952",
    #ELECTROMECANICA
    "2101", "2102", "2151", "2152", "2201", "2202", "2251", "2252", "2301", "2302", "2351", "2352", "2401", "2402", "2451", "2452", "2501", 
    "2502", "2551", "2552", "2601", "2602", "2651", "2652", "2751", "2752", "2851", "2852", "2951", "2952",
    #ADMINISTRACION
    "9101", "9102", "9151", "9152", "9201", "9202", "9251", "9252", "9301", "9302", "9351", "9352", "9401", "9402", "9451", "9452", "9501",
    "9502", "9551", "9552", "9601", "9602", "9651", "9652", "9751", "9752", "9851", "9852", "9951", "9952"
]

horarios = [
    "07:00", "08:00", "09:00", "10:00", "11:00", "12:00", "13:00",
    "14:00", "15:00", "16:00", "17:00", "18:00", "19:00", "20:00",
    "21:00"
]

carreras = [
    "ELECTRÓNICA", "INDUSTRIAL", "INFORMÁTICA",
    "ELECTROMECÁNICA", "SISTEMAS COMPUTACIONALES", "ADMINISTRACIÓN"
]

# Configuración de carga de archivos para la carrera de Ingeniería en Electrónica
ELECTRONICA_UPLOAD_FOLDER = "electronica/static/src/"
ELECTRONICA_ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

# Rutas CRUD para profesores en Ingeniería en Electrónica
@electronica_bp.route('/electronica_profesores', methods=['POST'])
@login_required('Electronica')
def add_profesor():
    data = request.json
    electronica_profesores.insert_one(data)
    return jsonify({'msg': 'Docente añadido'}), 201

@electronica_bp.route('/electronica_profesores/Add_Profesor')
@login_required('Electronica')
def add_profesor2():
    return render_template('electronica/Add_Profesor.html')

@electronica_bp.route('/electronica_profesores/<id>', methods=['GET'])
@login_required('Electronica')
def get_profesor(id):
    profesor = electronica_profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return jsonify(profesor)

@electronica_bp.route('/electronica_profesores/<id>', methods=['PUT'])
@login_required('Electronica')
def update_profesor(id):
    data = request.json
    electronica_profesores.update_one({'_id': ObjectId(id)}, {'$set': data})
    return jsonify({'msg': 'Docente actualizado'})

@electronica_bp.route('/electronica_profesores/edit/<id>', methods=['GET'])
@login_required('Electronica')
def edit_profesor(id):
    profesor = electronica_profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return render_template('electronica/edit_profesor.html', profesor=profesor, grupos=electronica_grupos, horarios=horarios)

@electronica_bp.route('/electronica_profesores/<id>', methods=['DELETE'])
@login_required('Electronica')
def delete_profesor(id):
    electronica_profesores.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Docente eliminado'})

@electronica_bp.route('/electronica_profesores', methods=['GET'])
@login_required('Electronica')
def get_all_profesores():
    all_profesores = list(electronica_profesores.find({
        "$or": [
            {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"}, {"carrera3": "ELECTRÓNICA"},
            {"carrera4": "ELECTRÓNICA"}, {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
            {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
            {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"}, {"carreraE3": "ELECTRÓNICA"},
            {"carreraE4": "ELECTRÓNICA"}, {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
            {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
            {"carreraC": "ELECTRÓNICA"}
        ]
    }).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return render_template('electronica/profesores.html', profesores=all_profesores)

# Ruta para obtener la lista de profesores en formato JSON en Ingeniería en Electrónica
@electronica_bp.route('/electronica_profesores/json', methods=['GET'])
@login_required('Electronica')
def get_profesores_json():
    all_profesores = list(electronica_profesores.find({
        "$or": [
            {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"}, {"carrera3": "ELECTRÓNICA"},
            {"carrera4": "ELECTRÓNICA"}, {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
            {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
            {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"}, {"carreraE3": "ELECTRÓNICA"},
            {"carreraE4": "ELECTRÓNICA"}, {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
            {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
            {"carreraC": "ELECTRÓNICA"}
        ]
    }, {'_id': 1, 'nombre': 1}).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return jsonify(all_profesores)

# Rutas CRUD para asignaturas en Ingeniería en Electrónica
@electronica_bp.route('/electronica_asignaturas', methods=['POST'])
@login_required('Electronica')
def add_asignatura():
    data = request.json
    electronica_asignaturas.insert_one(data)
    return jsonify({'msg': 'Asignatura añadida'}), 201

@electronica_bp.route('/electronica_asignaturas/<id>', methods=['GET'])
@login_required('Electronica')
def get_asignatura(id):
    asignatura = electronica_asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return jsonify(asignatura)

@electronica_bp.route('/electronica_asignaturas/<id>', methods=['PUT'])
@login_required('Electronica')
def update_asignatura(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = electronica_asignaturas.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Asignatura actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar la asignatura"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@electronica_bp.route('/electronica_edit-asignatura/<id>', methods=['GET'])
@login_required('Electronica')
def edit_asignatura(id):
    asignatura = electronica_asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return render_template('electronica/edit_asignatura.html', asignatura=asignatura)

@electronica_bp.route('/electronica_asignaturas/<id>', methods=['DELETE'])
@login_required('Electronica')
def delete_asignatura(id):
    electronica_asignaturas.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Asignatura eliminada'})

@electronica_bp.route('/electronica_asignaturas/json', methods=['GET'])
def get_all_asignaturas_json():
    all_asignaturas = list(electronica_asignaturas.find({}))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return jsonify(all_asignaturas)

@electronica_bp.route('/electronica_asignaturas', methods=['GET'])
@login_required('Electronica')
def get_all_asignaturas():
    all_asignaturas = list(electronica_asignaturas.find({}).sort("nombre", 1))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return render_template('electronica/asignaturas.html', asignaturas=all_asignaturas)

# Rutas CRUD para asignaturas especiales en Ingeniería en Electrónica
@electronica_bp.route('/electronica_asignaturasE', methods=['POST'])
@login_required('Electronica')
def add_asignaturaE():
    data = request.json
    electronica_asignaturasE.insert_one(data)
    return jsonify({'msg': 'Apoyo a la docencia añadida'}), 201

@electronica_bp.route('/electronica_asignaturasE/<id>', methods=['GET'])
@login_required('Electronica')
def get_asignaturaE(id):
    asignaturaE = electronica_asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(asignaturaE)

@electronica_bp.route('/electronica_asignaturasE/<id>', methods=['PUT'])
@login_required('Electronica')
def update_asignaturaE(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = electronica_asignaturasE.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Apoyo a la docencia actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar Apoyo a la docencia"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@electronica_bp.route('/electronica_edit-asignaturaE/<id>', methods=['GET'])
@login_required('Electronica')
def edit_asignaturaE(id):
    asignaturaE = electronica_asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('electronica/edit_asignaturaE.html', asignaturaE=asignaturaE)

@electronica_bp.route('/electronica_asignaturasE/<id>', methods=['DELETE'])
@login_required('Electronica')
def delete_asignaturaE(id):
    electronica_asignaturasE.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Apoyo a la docencia eliminada'})

@electronica_bp.route('/electronica_asignaturasE/json', methods=['GET'])
def get_all_asignaturasE_json():
    all_asignaturasE = list(electronica_asignaturasE.find({}))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(all_asignaturasE)

@electronica_bp.route('/electronica_asignaturasE', methods=['GET'])
@login_required('Electronica')
def get_all_asignaturasE():
    all_asignaturasE = list(electronica_asignaturasE.find({}).sort("nombre", 1))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('electronica/asignaturasE.html', asignaturasE=all_asignaturasE)

# Función para exportar datos en Ingeniería en Electrónica
# Rutas para obtener las columnas disponibles
@electronica_bp.route('/electronica_columns/<collection_name>', methods=['GET'])
@login_required('Electronica')
def get_columns(collection_name):
    column_mappings = {
        "profesores": [
            "nombre", "profesion", "adscripcion", "fecha_ingreso",
            "tiempo_indeterminado", "periodo_actual", "horas_a",
            "horas_b", "Horas de Asignatura", "Horas Descarga", "total_horas"
        ],
        "asignaturas": ["nombre", "horas"],
        "asignaturasE": ["nombre", "horas"]  # Nueva colección de Asignaturas Especiales
    }

    if collection_name not in column_mappings:
        return jsonify({"error": "Colección no encontrada"}), 404

    return jsonify(column_mappings[collection_name])

# Función para exportar datos
@electronica_bp.route('/electronica_export', methods=['POST'])
@login_required('Electronica')
def export_data():
    data = request.json
    selected_columns = data['columns'].split(',')
    collection_name = data['collection']
    export_format = data.get('format', 'xlsx')
    
     # Seleccionar la colección correcta
    if collection_name == "profesores":
        collection = electronica_profesores
    elif collection_name == "asignaturas":
        collection = electronica_asignaturas
    elif collection_name == "asignaturasE":
        collection = electronica_asignaturasE
    else:
        return jsonify({"error": "Colección no válida"}), 400

    # Filtrar los datos según la colección seleccionada
    if collection_name == "profesores":
        cursor = collection.find({
            "$or": [
                {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"}, {"carrera3": "ELECTRÓNICA"},
                {"carrera4": "ELECTRÓNICA"}, {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
                {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
                {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"}, {"carreraE3": "ELECTRÓNICA"},
                {"carreraE4": "ELECTRÓNICA"}, {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
                {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
                {"carreraC": "ELECTRÓNICA"}
            ]
        })
    else:
        cursor = collection.find({})  # Para asignaturas y asignaturasE, no aplicamos filtros específicos

    df = pd.DataFrame(list(cursor))

    # Reemplazar valores NaN e Inf en el DataFrame antes de exportarlo
    df = df.replace([float('inf'), -float('inf')], 0).fillna('')

    if df.empty:
        return {"message": "No hay datos para exportar"}, 400

    # Ordenar alfabéticamente por "nombre"
    df = df.sort_values(by="nombre", ascending=True)

    # **📌 Si el usuario selecciona estas columnas, las calculamos**
    if "Horas de Asignatura" in selected_columns:
        df["Horas de Asignatura"] = df.apply(lambda row: sum(
            int(row.get(f"horas{i}", 0) or 0) for i in range(1, 9) if row.get(f"carrera{i}") == "ELECTRÓNICA"
        ), axis=1)

    if "Horas Descarga" in selected_columns:
        df["Horas Descarga"] = df.apply(lambda row: sum(
            int(row.get(f"horasE{i}", 0) or 0) for i in range(1, 9) if row.get(f"carreraE{i}") == "ELECTRÓNICA"
        ), axis=1)

    if collection_name == "profesores":
        output = BytesIO()
        if export_format == 'xlsx':
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book
                sheet = workbook.add_worksheet('Datos')

                # Definir colores alternos
                format_white = workbook.add_format({'bg_color': '#FFFFFF'})  # Blanco
                format_gray = workbook.add_format({'bg_color': '#F2F2F2'})  # Gris claro
                format_header = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF'})
                format_red = workbook.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})  # Rojo para errores

                # **📌 Filtrar solo las columnas seleccionadas**
                selected_columns_filtered = [col for col in selected_columns if col not in [
                    "asignacion_horas_frente_grupo", "asignacion_horas_descarga_otras_actividades", "asignacion_horas_cargo_academico"
                ]]

                # Escribir encabezados
                sheet.write_row(0, 0, selected_columns_filtered, format_header)

                # Escribir datos con colores alternos
                for row_num, row in enumerate(df[selected_columns_filtered].values, start=1):
                    color_format = format_gray if row_num % 2 == 0 else format_white

                    # Escribir toda la fila normalmente
                    sheet.write_row(row_num, 0, row, color_format)

                    # Verificar si "total_horas" debe ser pintado de rojo
                    if "total_horas" in selected_columns_filtered:
                        total_horas_index = selected_columns_filtered.index("total_horas")
                        total_horas = int(row[total_horas_index]) if row[total_horas_index] else 0
                        horas_asignatura = int(row[selected_columns_filtered.index("Horas de Asignatura")]) if "Horas de Asignatura" in selected_columns_filtered else 0
                        horas_descarga = int(row[selected_columns_filtered.index("Horas Descarga")]) if "Horas Descarga" in selected_columns_filtered else 0

                        # Verificar si la diferencia es distinta de 0
                        if (horas_asignatura + horas_descarga) != total_horas:
                            sheet.write(row_num, selected_columns_filtered.index("total_horas"), total_horas, format_red)

                # **📌 1. Horas Frente a Grupo**
                if 'asignacion_horas_frente_grupo' in selected_columns:
                    sheet_horarios = workbook.add_worksheet("Horas Frente a Grupo")
                    headers = ["Carrera", "Asignatura", "Grupo", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                    sheet_horarios.write_row(0, 0, headers, format_header)

                    row_num = 1
                    for index, profesor in enumerate(df.to_dict(orient="records")):
                        color_format = format_gray if index % 2 == 0 else format_white
                        for i in range(1, 9):
                            sheet_horarios.write(row_num, 0, profesor.get(f"carrera{i}", ""), color_format)
                            sheet_horarios.write(row_num, 1, profesor.get(f"asignatura{i}", ""), color_format)
                            sheet_horarios.write(row_num, 2, profesor.get(f"grupo{i}", ""), color_format)
                            sheet_horarios.write(row_num, 3, profesor.get(f"horas{i}", 0), color_format)

                            # Horarios por día (Lunes - Sábado)
                            for j in range(1, 7):  
                                horario_inicio = profesor.get(f"hora_inicio{i}{j}", "")
                                horario_fin = profesor.get(f"hora_fin{i}{j}", "")
                                sheet_horarios.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                            row_num += 1

                # **📌 2. Horas Descarga Otras Actividades**
                if 'asignacion_horas_descarga_otras_actividades' in selected_columns:
                    sheet_descarga = workbook.add_worksheet("Horas Descarga")
                    headers_descarga = ["Carrera", "Asignatura", "Grupo", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                    sheet_descarga.write_row(0, 0, headers_descarga, format_header)

                    row_num = 1
                    for index, profesor in enumerate(df.to_dict(orient="records")):
                        color_format = format_gray if index % 2 == 0 else format_white
                        for i in range(1, 9):
                            sheet_descarga.write(row_num, 0, profesor.get(f"carreraE{i}", ""), color_format)
                            sheet_descarga.write(row_num, 1, profesor.get(f"asignaturaE{i}", ""), color_format)
                            sheet_descarga.write(row_num, 2, profesor.get(f"grupoE{i}", ""), color_format)
                            sheet_descarga.write(row_num, 3, profesor.get(f"horasE{i}", 0), color_format)

                            for j in range(1, 7):
                                horario_inicio = profesor.get(f"hora_inicioE{i}{j}", "")
                                horario_fin = profesor.get(f"hora_finE{i}{j}", "")
                                sheet_descarga.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                            row_num += 1

                # **📌 3. Horas Cargo Académico**
                if 'asignacion_horas_cargo_academico' in selected_columns:
                    sheet_cargo = workbook.add_worksheet("Horas Cargo Académico")
                    headers_cargo = ["Carrera", "Cargo", "Vigencia", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
                    sheet_cargo.write_row(0, 0, headers_cargo, format_header)

                    row_num = 1
                    for index, profesor in enumerate(df.to_dict(orient="records")):
                        color_format = format_gray if index % 2 == 0 else format_white
                        sheet_cargo.write(row_num, 0, profesor.get("carreraC", ""), color_format)  # Primera carrera registrada
                        sheet_cargo.write(row_num, 1, profesor.get("cargo", ""), color_format)
                        sheet_cargo.write(row_num, 2, profesor.get("vigenciaCargo", ""), color_format)
                        sheet_cargo.write(row_num, 3, profesor.get("horasC", 0), color_format)

                        for j in range(1, 7):
                            horario_inicio = profesor.get(f"hora_inicioC1{j}", "")
                            horario_fin = profesor.get(f"hora_finC1{j}", "")
                            sheet_cargo.write(row_num, 3 + j, f"{horario_inicio} - {horario_fin}", color_format)

                        row_num += 1

            output.seek(0)
            return send_file(output, as_attachment=True, download_name=f"{collection_name}.xlsx")
    else:
        output = BytesIO()
        if export_format == 'xlsx':
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book
                sheet = workbook.add_worksheet('Datos')

                # 📌 Definir formatos
                format_white = workbook.add_format({'bg_color': '#FFFFFF'})
                format_gray = workbook.add_format({'bg_color': '#F2F2F2'})
                format_header = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF'})

                # 📌 Filtrar solo las columnas seleccionadas
                selected_columns_filtered = [col for col in selected_columns]

                # 📌 Escribir encabezados
                sheet.write_row(0, 0, selected_columns_filtered, format_header)

                # 📌 Escribir datos con colores alternos
                for row_num, row in enumerate(df[selected_columns_filtered].values, start=1):
                    color_format = format_gray if row_num % 2 == 0 else format_white
                    sheet.write_row(row_num, 0, row, color_format)

            output.seek(0)
            return send_file(output, as_attachment=True, download_name=f"{collection_name}.xlsx")

#RUTAS PARA MOSTRAR Y DESCARGAR ARCHIVOS AUTOMATICOS
@electronica_bp.route('/historial')
@login_required('Electronica')
def historial_exportaciones():
    """Muestra la lista de archivos exportados almacenados en GridFS para Sistemas Computacionales."""
    files = list(db.fs.files.find(
        {"filename": {"$regex": "^Datos_Electronica_"}},
        {"filename": 1, "uploadDate": 1}
    ).sort("uploadDate", -1))  # Orden descendente por fecha
    return render_template("electronica/historial.html", files=files)

@electronica_bp.route('/historial/download/<filename>')
@login_required('Electronica')
def download_file(filename):
    file = db.fs.files.find_one({"filename": filename})

    if not file:
        return "Archivo no encontrado", 404

    file_data = fs.get(file["_id"])  # Obtener el archivo desde GridFS

    # Enviar archivo como respuesta para descarga
    return Response(
        file_data.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

# 📌 Ruta para eliminar archivos de GridFS en ELECTRONICA
@electronica_bp.route('/historial/delete/<filename>', methods=['DELETE'])
@login_required('Electronica')
def delete_file(filename):
    """Elimina un archivo de GridFS correspondiente a Sistemas Computacionales"""
    try:
        # Buscar el archivo en GridFS
        file_to_delete = fs.find_one({"filename": filename})

        if file_to_delete:
            fs.delete(file_to_delete._id)  # Eliminar el archivo de GridFS
            return jsonify({"success": True, "message": "Archivo eliminado correctamente."})
        else:
            return jsonify({"success": False, "message": "Archivo no encontrado."}), 404

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

#EXPORTACION DE DATOS AUTOMATICA
def export_data_auto():
    """Ejecuta la exportación automática el 10 de junio y 25 de octubre."""

    current_date = datetime.now()
    current_year = current_date.year

    # 📌 Definir el semestre basado en la fecha actual
    if 2 <= current_date.month <= 7:
        period = "1"
    elif 8 <= current_date.month <= 1:
        period = "2"
    else:
        print("📌 No es una fecha de exportación automática. Se cancela la ejecución.")
        return
    
    # 📌 Nombre del archivo
    filename = f"Datos_Electronica_{current_year}-{period}.xlsx"

    # 📌 Eliminar archivo previo si ya existe en GridFS
    #existing_file = db.fs.files.find_one({"filename": filename})
    #if existing_file:
    #    fs.delete(existing_file["_id"])
    #    print(f"📌 Archivo previo {filename} eliminado de GridFS.")

    # 📌 Columnas a exportar
    selected_columns = [
        "nombre", "profesion", "adscripcion", "fecha_ingreso",
        "tiempo_indeterminado", "periodo_actual", "horas_a",
        "horas_b", "Horas de Asignatura", "Horas Descarga", "total_horas",
        "asignacion_horas_frente_grupo", "asignacion_horas_descarga_otras_actividades",
        "asignacion_horas_cargo_academico"
    ]

    # 📌 Seleccionar colección específica de la carrera
    collection_name = "profesores"
    collection = db[collection_name]

    # Filtrar solo profesores que pertenezcan a Industrial
    cursor = collection.find({
        "$or": [
            {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"},
            {"carrera3": "ELECTRÓNICA"}, {"carrera4": "ELECTRÓNICA"},
            {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
            {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
            {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"},
            {"carreraE3": "ELECTRÓNICA"}, {"carreraE4": "ELECTRÓNICA"},
            {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
            {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
            {"carreraC": "ELECTRÓNICA"}
        ]
    })
    df = pd.DataFrame(list(cursor))

    if df.empty:
        print("⚠ No hay datos para exportar.")
        return

    df = df.replace([float('inf'), -float('inf')], 0).fillna('')
    df = df.sort_values(by="nombre", ascending=True)

    # 📌 Asegurar que todas las columnas existan en `df`
    for col in selected_columns:
        if col not in df.columns:
            df[col] = 0 if "horas" in col.lower() else ""

    # 📌 Calcular Horas de Asignatura y Horas Descarga
    df["Horas de Asignatura"] = df.apply(lambda row: sum(
        int(row.get(f"horas{i}", 0) or 0) for i in range(1, 9) if row.get(f"carrera{i}") == "INDUSTRIAL"
    ), axis=1)

    df["Horas Descarga"] = df.apply(lambda row: sum(
        int(row.get(f"horasE{i}", 0) or 0) for i in range(1, 9) if row.get(f"carreraE{i}") == "INDUSTRIAL"
    ), axis=1)

    # 📌 Crear un archivo en memoria para almacenar el Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        sheet = workbook.add_worksheet('Datos')

        # 📌 Definir formatos de celda
        format_white = workbook.add_format({'bg_color': '#FFFFFF'})
        format_gray = workbook.add_format({'bg_color': '#F2F2F2'})
        format_header = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF'})
        format_red = workbook.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})

        # 📌 Escribir encabezados
        sheet.write_row(0, 0, selected_columns, format_header)

        # 📌 Escribir datos con colores alternos
        for row_num, row in enumerate(df[selected_columns].values, start=1):
            color_format = format_gray if row_num % 2 == 0 else format_white
            sheet.write_row(row_num, 0, row, color_format)

            # 📌 Validar `total_horas`
            total_horas = int(row[selected_columns.index("total_horas")]) if row[selected_columns.index("total_horas")] else 0
            horas_asignatura = int(row[selected_columns.index("Horas de Asignatura")]) if "Horas de Asignatura" in selected_columns else 0
            horas_descarga = int(row[selected_columns.index("Horas Descarga")]) if "Horas Descarga" in selected_columns else 0

            if (horas_asignatura + horas_descarga) != total_horas:
                sheet.write(row_num, selected_columns.index("total_horas"), total_horas, format_red)

        # 📌 1. Horas Frente a Grupo
        sheet_horarios = workbook.add_worksheet("Horas Frente a Grupo")
        headers = ["Carrera", "Asignatura", "Grupo", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
        sheet_horarios.write_row(0, 0, headers, format_header)

        row_num = 1
        for index, profesor in enumerate(df.to_dict(orient="records")):
            color_format = format_gray if index % 2 == 0 else format_white
            for i in range(1, 9):
                sheet_horarios.write_row(row_num, 0, [
                    profesor.get(f"carrera{i}", ""), profesor.get(f"asignatura{i}", ""),
                    profesor.get(f"grupo{i}", ""), profesor.get(f"horas{i}", 0)
                ] + [f"{profesor.get(f'hora_inicio{i}{j}', '')} - {profesor.get(f'hora_fin{i}{j}', '')}" for j in range(1, 7)],
                color_format)
                row_num += 1

        # 📌 2. Horas Descarga Otras Actividades
        sheet_descarga = workbook.add_worksheet("Horas Descarga")
        sheet_descarga.write_row(0, 0, headers, format_header)

        row_num = 1
        for index, profesor in enumerate(df.to_dict(orient="records")):
            color_format = format_gray if index % 2 == 0 else format_white
            for i in range(1, 9):
                sheet_descarga.write_row(row_num, 0, [
                    profesor.get(f"carreraE{i}", ""), profesor.get(f"asignaturaE{i}", ""),
                    profesor.get(f"grupoE{i}", ""), profesor.get(f"horasE{i}", 0)
                ] + [f"{profesor.get(f'hora_inicioE{i}{j}', '')} - {profesor.get(f'hora_finE{i}{j}', '')}" for j in range(1, 7)],
                color_format)
                row_num += 1

        # 📌 3. Horas Cargo Académico
        sheet_cargo = workbook.add_worksheet("Horas Cargo Académico")
        sheet_cargo.write_row(0, 0, ["Carrera", "Cargo", "Vigencia", "Horas", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"], format_header)

        row_num = 1
        for index, profesor in enumerate(df.to_dict(orient="records")):
            color_format = format_gray if index % 2 == 0 else format_white
            sheet_cargo.write_row(row_num, 0, [
                profesor.get("carreraC", ""), profesor.get("cargo", ""), profesor.get("vigenciaCargo", ""),
                profesor.get("horasC", 0)
            ] + [f"{profesor.get(f'hora_inicioC1{j}', '')} - {profesor.get(f'hora_finC1{j}', '')}" for j in range(1, 7)],
            color_format)
            row_num += 1

    # 📌 Guardar el archivo en GridFS
    output.seek(0)
    file_id = fs.put(output.read(), filename=filename, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    print(f"✅ Exportación automática completada y guardada en GridFS con ID: {file_id}")

@electronica_bp.route("/exportar-manual", methods=["POST"])
@login_required("Electronica")
def exportar_manual():
    export_data_auto()
    return jsonify({"mensaje": f"Exportación manual realizada con éxito"})

# Rutas de la interfaz de usuario para Ingeniería en Electrónica
@electronica_bp.route('/electronica_index')
@login_required('Electronica')
def index():
    return render_template('electronica/index.html')

@electronica_bp.route('/electronica')
@login_required('Electronica')
def principal():
    return render_template('electronica/principal.html')

@electronica_bp.route('/electronica_reporteador')
@login_required('Electronica')
def reporteador():
    return render_template('electronica/exportacion/exportar.html')

# Validar archivos permitidos para imágenes
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# Ruta para subir imágenes de encabezado y pie de página en Electrónica
@electronica_bp.route("/electronica_upload-images", methods=["POST"])
@login_required('Electronica')
def upload_images():
    upload_folder = "static/electronica/src/"

    if "header" in request.files:
        header = request.files["header"]
        if header and allowed_file(header.filename):
            header.save(os.path.join(upload_folder, "Encabezado1.PNG"))

    if "footer" in request.files:
        footer = request.files["footer"]
        if footer and allowed_file(footer.filename):
            footer.save(os.path.join(upload_folder, "PieDePagina1.PNG"))

    return jsonify({'msg': 'Imágenes actualizadas correctamente'}), 200

# Ruta para actualizar el texto en la celda A4 en Electrónica
@electronica_bp.route('/electronica_update-text', methods=['POST'])
@login_required('Electronica')
def update_text():
    nuevo_texto = request.form.get('nuevo_texto', '')
    if not nuevo_texto:
        return jsonify({'msg': 'Error: No se proporcionó un texto válido'}), 400
    
    # Guardar el nuevo texto en un archivo de configuración
    texto_a4_path = "static/electronica/src/texto_a4.txt"
    with open(texto_a4_path, "w", encoding="utf-8") as file:
        file.write(nuevo_texto)
    
    return jsonify({'msg': 'Texto A4 actualizado correctamente'}), 200

# Ruta para actualizar el texto en la celda A54 en Electrónica
@electronica_bp.route('/electronica_update-text-a54', methods=['POST'])
@login_required('Electronica')
def update_text_a54():
    nuevo_texto_dos = request.form.get('nuevo_texto_dos', '')
    if not nuevo_texto_dos:
        return jsonify({'msg': 'Error: No se proporcionó un texto válido'}), 400
    
    # Guardar el nuevo texto en un archivo de configuración
    texto_a54_path = "static/electronica/src/texto_a54.txt"
    with open(texto_a54_path, "w", encoding="utf-8") as file:
        file.write(nuevo_texto_dos)
    
    return jsonify({'msg': 'Texto A54 actualizado correctamente'}), 200

# Ruta para exportar los profesores seleccionados usando la plantilla con 32 hojas EXCEL en Electrónica
@electronica_bp.route('/electronica_export-selected', methods=['POST'])
@login_required('Electronica')
def export_selected():
    profesor_ids = request.form.getlist('profesor_ids')
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [electronica_profesores.find_one({'_id': ObjectId(profesor_id),
        "$or": [
            {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"}, {"carrera3": "ELECTRÓNICA"},
            {"carrera4": "ELECTRÓNICA"}, {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
            {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
            {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"}, {"carreraE3": "ELECTRÓNICA"},
            {"carreraE4": "ELECTRÓNICA"}, {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
            {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
            {"carreraC": "ELECTRÓNICA"}
        ]}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas
    template_path = "static/electronica/src/Plantilla_pie_reducido_2cm.xlsx"
    workbook = openpyxl.load_workbook(template_path)

    # Establecer bordes
    thin_side = Side(style='thin')
    border_A44 = Border(bottom=thin_side, left=thin_side)
    border_B44 = Border(bottom=thin_side)

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 y A54 desde archivos de configuración
    texto_a4_path = "static/electronica/src/texto_a4.txt"
    texto_a54_path = "static/electronica/src/texto_a54.txt"
    
    texto_a4 = open(texto_a4_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a4_path) else ""
    texto_a54 = open(texto_a54_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a54_path) else ""

    # Obtener el nombre y cargo del encargado de dirección académica y general
    encargado = administrativos.find_one({'cargo': {'$in': ["ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "DIRECTOR ACADÉMICO", "DIRECTORA ACADÉMICA"]}})
    direccion = administrativos.find_one({'cargo': {'$in': ["DIRECTOR GENERAL", "DIRECTORA GENERAL"]}})

    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos en Electrónica
    cargo_mapping = {
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "ELECTROMECÁNICA": ["JEFA DE DIVISIÓN DE ING. ELECTROMECÁNICA", "JEFE DE DIVISIÓN DE ING. ELECTROMECÁNICA"],
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        # ⚠️ IMPORTANTE: Crear nuevas instancias de la imagen en cada iteración
        header_image_copy = Image("static/electronica/src/Encabezado1.PNG")
        footer_image_copy = Image("static/electronica/src/PieDePagina1.PNG")

        # Insertar imágenes en encabezado y pie de página
        sheet.add_image(header_image_copy, "A1")
        sheet.add_image(footer_image_copy, "A56")

        # Actualizar el texto en la celda A4 y A54
        sheet["A4"] = texto_a4
        sheet["A54"] = texto_a54

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de dirección
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("2"):
                    sheet[f"A{row}"] = "ELECTROMECÁNICA"
                elif str(grupo).startswith("3"):
                    sheet[f"A{row}"] = "ELECTRÓNICA"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"
                elif str(grupo).startswith("6"):
                    sheet[f"A{row}"] = "INFORMÁTICA"
                elif str(grupo).startswith("9"):
                    sheet[f"A{row}"] = "ADMINISTRACIÓN"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)

        # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado1 = encargado["nombre"] if encargado else ""
            cargo_encargado1 = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado1
            sheet["A46"] = nombre_encargado1
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_A44
            sheet["B44"].border = border_B44

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # Guardar el archivo en memoria
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Devolver el archivo actualizado como respuesta
    return send_file(
        output,
        as_attachment=True,
        download_name="Reporte_Profesores_Electronica.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 📌 Función mejorada para convertir Excel a PDF usando LibreOffice
def excel_to_pdf_libreoffice(input_excel_path, output_pdf_path):
    try:
        # 📌 Verifica que el archivo Excel existe antes de la conversión
        if not os.path.exists(input_excel_path):
            print(f"❌ Error: El archivo Excel {input_excel_path} no existe.")
            return False

        libreoffice_path = "C:\\Program Files\\LibreOffice\\program\\soffice.exe"  # Ajustar según instalación

        # 📌 Comando para convertir Excel a PDF
        cmd = [
            libreoffice_path, "--headless", "--convert-to", "pdf",
            input_excel_path, "--outdir", os.path.dirname(output_pdf_path)
        ]

        print(f"📄 Ejecutando comando: {' '.join(cmd)}")  # Debug: Ver el comando

        # 📌 Ejecutar el comando y capturar la salida
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        # 📌 Mostrar la salida del comando en caso de error
        if result.returncode != 0:
            print("❌ Error al convertir Excel a PDF:", result.stderr)
            return False

        # 📌 Asegurar que el PDF fue generado
        generated_pdf = input_excel_path.replace(".xlsx", ".pdf")  # LibreOffice cambia automáticamente la extensión
        if not os.path.exists(generated_pdf):
            print("❌ Error: No se generó el archivo PDF.")
            return False

        # 📌 Eliminar el archivo PDF de salida si ya existe
        if os.path.exists(output_pdf_path):
            os.remove(output_pdf_path)
            print(f"🔄 Archivo existente {output_pdf_path} eliminado antes de renombrar.")

        # 📌 Renombrar el PDF al nombre esperado
        os.rename(generated_pdf, output_pdf_path)
        print(f"✅ Conversión exitosa: {output_pdf_path}")
        return True

    except Exception as e:
        print("❌ Error inesperado al convertir Excel a PDF:", e)
        return False

# Ruta para exportar los profesores seleccionados en PDF en Electrónica
@electronica_bp.route('/electronica_export-selected-pdf', methods=['POST'])
@login_required('Electronica')
def export_selected_pdf_electronica():
    profesor_ids = request.form.getlist('profesor_ids')

    if not profesor_ids:
        return "Error: No se recibieron IDs de profesores", 400
    
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [electronica_profesores.find_one({'_id': ObjectId(profesor_id),
        "$or": [
            {"carrera1": "ELECTRÓNICA"}, {"carrera2": "ELECTRÓNICA"}, {"carrera3": "ELECTRÓNICA"},
            {"carrera4": "ELECTRÓNICA"}, {"carrera5": "ELECTRÓNICA"}, {"carrera6": "ELECTRÓNICA"},
            {"carrera7": "ELECTRÓNICA"}, {"carrera8": "ELECTRÓNICA"},
            {"carreraE1": "ELECTRÓNICA"}, {"carreraE2": "ELECTRÓNICA"}, {"carreraE3": "ELECTRÓNICA"},
            {"carreraE4": "ELECTRÓNICA"}, {"carreraE5": "ELECTRÓNICA"}, {"carreraE6": "ELECTRÓNICA"},
            {"carreraE7": "ELECTRÓNICA"}, {"carreraE8": "ELECTRÓNICA"},
            {"carreraC": "ELECTRÓNICA"}
        ]}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas en Electrónica
    template_path = "static/electronica/src/Plantilla_pie_reducido_2cm.xlsx"
    workbook = openpyxl.load_workbook(template_path)

    # Establecer bordes
    thin_side = Side(style='thin')
    border_A44 = Border(bottom=thin_side, left=thin_side)
    border_B44 = Border(bottom=thin_side)

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 y A54 desde archivos de configuración
    texto_a4_path = "static/electronica/src/texto_a4.txt"
    texto_a54_path = "static/electronica/src/texto_a54.txt"
    
    texto_a4 = open(texto_a4_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a4_path) else ""
    texto_a54 = open(texto_a54_path, "r", encoding="utf-8").read().strip() if os.path.exists(texto_a54_path) else ""

    # Obtener el nombre y cargo del encargado de dirección académica y general
    encargado = administrativos.find_one({'cargo': {'$in': ["ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "DIRECTOR ACADÉMICO", "DIRECTORA ACADÉMICA"]}})
    direccion = administrativos.find_one({'cargo': {'$in': ["DIRECTOR GENERAL", "DIRECTORA GENERAL"]}})

    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos en Electrónica
    cargo_mapping = {
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "ELECTROMECÁNICA": ["JEFA DE DIVISIÓN DE ING. ELECTROMECÁNICA", "JEFE DE DIVISIÓN DE ING. ELECTROMECÁNICA"],
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. EN SISTEMAS COMPUTACIONALES"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        # ⚠️ IMPORTANTE: Crear nuevas instancias de la imagen en cada iteración
        header_image_copy = Image("static/electronica/src/Encabezado1.PNG")
        footer_image_copy = Image("static/electronica/src/PieDePagina1.PNG")

        # Insertar imágenes en encabezado y pie de página
        sheet.add_image(header_image_copy, "A1")
        sheet.add_image(footer_image_copy, "A56")

        # Actualizar el texto en la celda A4 y A54
        sheet["A4"] = texto_a4
        sheet["A54"] = texto_a54

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de dirección
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("2"):
                    sheet[f"A{row}"] = "ELECTROMECÁNICA"
                elif str(grupo).startswith("3"):
                    sheet[f"A{row}"] = "ELECTRÓNICA"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"
                elif str(grupo).startswith("6"):
                    sheet[f"A{row}"] = "INFORMÁTICA"
                elif str(grupo).startswith("9"):
                    sheet[f"A{row}"] = "ADMINISTRACIÓN"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)

        # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado1 = encargado["nombre"] if encargado else ""
            cargo_encargado1 = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado1
            sheet["A46"] = nombre_encargado1
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_A44
            sheet["B44"].border = border_B44

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # 📌 Guardar el archivo Excel antes de convertirlo a PDF
    temp_excel_path = "temp_reporte_Electronica.xlsx"
    pdf_path = "Reporte_Profesores_Electronica.pdf"
    
    workbook.save(temp_excel_path)
    workbook.close()

    # 📌 Esperar un breve momento para garantizar que el archivo se guarde correctamente
    time.sleep(2)

    # 📌 Convertir Excel a PDF
    conversion_exitosa = excel_to_pdf_libreoffice(temp_excel_path, pdf_path)

    if not conversion_exitosa:
        return "Error al generar el PDF", 500

    # 📌 Enviar el archivo PDF generado
    return send_file(
        pdf_path,
        as_attachment=True,
        download_name="Reporte_Profesores_Electronica.pdf",
        mimetype="application/pdf"
    )