from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
from io import BytesIO
from io import StringIO
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font
import win32com.client
import os
import pythoncom
from datetime import datetime

app = Flask(__name__)

client = MongoClient('mongodb+srv://ivan:tuliogaymer077@cluster0.bkahq7u.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0')
db = client.tecnologico

profesores = db['profesores']
asignaturas = db['asignaturas']
asignaturasE = db['asignaturasE']
administrativos = db['administrativos']

grupos = ["NG","1101", "1102", "1151", "1152", "1181", "1201", "1202", "1251", "1252", 
          "1281", "1301", "1302", "1351", "1352", "1381", "1401", "1402", "1451", 
          "1452", "1481", "1501", "1502", "1551", "1552", "1581", "1601", "1602", 
          "1651", "1652", "1681", "1751", "1752", "1781", "1851", "1852", "1881", 
          "1951", "1952", "1981"]

horarios = ["07:00", "08:00", "09:00", "10:00", "11:00", "12:00", "13:00", 
            "14:00", "15:00", "16:00", "17:00", "18:00", "19:00", "20:00", 
            "21:00"]

carreras = ["INDUSTRIAL", "SISTEMAS COMPUTACIONALES", "ELECTRÓNICA",
            "MECATRÓNICA", "INFORMÁTICA", "ADMINISTRACIÓN"]

UPLOAD_FOLDER = "static/src/"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Rutas CRUD para profesores
@app.route('/profesores', methods=['POST'])
def add_profesor():
    data = request.json
    profesores.insert_one(data)
    return jsonify({'msg': 'Profesor añadido'}), 201

@app.route('/Add_Profesor')
def add_profesor2():
    return render_template('Add_Profesor.html')

@app.route('/profesores/<id>', methods=['GET'])
def get_profesor(id):
    profesor = profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return jsonify(profesor)

@app.route('/profesores/<id>', methods=['PUT'])
def update_profesor(id):
    data = request.json
    profesores.update_one({'_id': ObjectId(id)}, {'$set': data})
    return jsonify({'msg': 'Profesor actualizado'})

@app.route('/edit-profesor/<id>', methods=['GET'])
def edit_profesor(id):
    profesor = profesores.find_one({'_id': ObjectId(id)})
    profesor['_id'] = str(profesor['_id'])
    return render_template('edit_profesor.html', profesor=profesor, grupos=grupos, horarios=horarios)

@app.route('/profesores/<id>', methods=['DELETE'])
def delete_profesor(id):
    profesores.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Profesor eliminado'})

@app.route('/profesores', methods=['GET'])
def get_all_profesores():
    all_profesores = list(profesores.find({}).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return render_template('profesores.html', profesores=all_profesores)

# Ruta para obtener la lista de profesores en formato JSON
@app.route('/profesores/json', methods=['GET'])
def get_profesores_json():
    all_profesores = list(profesores.find({}, {'_id': 1, 'nombre': 1}).sort("nombre", 1))
    for profesor in all_profesores:
        profesor['_id'] = str(profesor['_id'])
    return jsonify(all_profesores)

# Rutas CRUD para asignaturas
@app.route('/asignaturas', methods=['POST'])
def add_asignatura():
    data = request.json
    asignaturas.insert_one(data)
    return jsonify({'msg': 'Asignatura añadida'}), 201

@app.route('/asignaturas/<id>', methods=['GET'])
def get_asignatura(id):
    asignatura = asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return jsonify(asignatura)

@app.route('/asignaturas/<id>', methods=['PUT'])
def update_asignatura(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = asignaturas.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Asignatura actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar la asignatura"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@app.route('/edit-asignatura/<id>', methods=['GET'])
def edit_asignatura(id):
    asignatura = asignaturas.find_one({'_id': ObjectId(id)})
    asignatura['_id'] = str(asignatura['_id'])
    return render_template('edit_asignatura.html', asignatura=asignatura)

@app.route('/asignaturas/<id>', methods=['DELETE'])
def delete_asignatura(id):
    asignaturas.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Asignatura eliminada'})

@app.route('/asignaturas/json', methods=['GET'])
def get_all_asignaturas_json():
    all_asignaturas = list(asignaturas.find({}))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return jsonify(all_asignaturas)

# Ruta para obtener todas las asignaturas y renderizar el template
@app.route('/asignaturas', methods=['GET'])
def get_all_asignaturas():
    all_asignaturas = list(asignaturas.find({}).sort("nombre", 1))
    for asignatura in all_asignaturas:
        asignatura['_id'] = str(asignatura['_id'])
    return render_template('asignaturas.html', asignaturas=all_asignaturas)

# Rutas CRUD para asignaturas especiales
@app.route('/asignaturasE', methods=['POST'])
def add_asignaturaE():
    data = request.json
    asignaturasE.insert_one(data)
    return jsonify({'msg': 'Asignatura Especial añadida'}), 201

@app.route('/asignaturasE/<id>', methods=['GET'])
def get_asignaturaE(id):
    asignaturaE = asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(asignaturaE)

@app.route('/asignaturasE/<id>', methods=['PUT'])
def update_asignaturaE(id):
    data = request.json
    nombre = data.get('nombre')
    horas = data.get('horas')

    if nombre and isinstance(horas, int):
        result = asignaturasE.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'horas': horas}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Asignatura Especial actualizada"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar la Asignatura Especial"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@app.route('/edit-asignaturaE/<id>', methods=['GET'])
def edit_asignaturaE(id):
    asignaturaE = asignaturasE.find_one({'_id': ObjectId(id)})
    asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('edit_asignaturaE.html', asignaturaE=asignaturaE)

@app.route('/asignaturasE/<id>', methods=['DELETE'])
def delete_asignaturaE(id):
    asignaturasE.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Asignatura Especial eliminada'})

@app.route('/asignaturasE/json', methods=['GET'])
def get_all_asignaturasE_json():
    all_asignaturasE = list(asignaturasE.find({}))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return jsonify(all_asignaturasE)

# Ruta para obtener todas las asignaturas y renderizar el template
@app.route('/asignaturasE', methods=['GET'])
def get_all_asignaturasE():
    all_asignaturasE = list(asignaturasE.find({}).sort("nombre", 1))
    for asignaturaE in all_asignaturasE:
        asignaturaE['_id'] = str(asignaturaE['_id'])
    return render_template('asignaturasE.html', asignaturasE=all_asignaturasE)

# Rutas CRUD para administrativos
@app.route('/administrativos', methods=['POST'])
def add_administrativo():
    data = request.json
    administrativos.insert_one(data)
    return jsonify({'msg': 'Administrativo añadido'}), 201

@app.route('/administrativos/<id>', methods=['GET'])
def get_administrativo(id):
    administrativo = administrativos.find_one({'_id': ObjectId(id)})
    administrativo['_id'] = str(administrativo['_id'])
    return jsonify(administrativo)

@app.route('/administrativos/<id>', methods=['PUT'])
def update_administrativo(id):
    data = request.json
    nombre = data.get('nombre')
    cargo = data.get('cargo')

    if nombre and cargo:
        result = administrativos.update_one(
            {'_id': ObjectId(id)},
            {'$set': {'nombre': nombre, 'cargo': cargo}}
        )

        if result.modified_count > 0:
            return jsonify({"msg": "Administrativo actualizado"}), 200
        else:
            return jsonify({"msg": "No se pudo actualizar el Administrativo"}), 400
    else:
        return jsonify({"msg": "Datos inválidos"}), 400

@app.route('/edit-administrativo/<id>', methods=['GET'])
def edit_administrativo(id):
    administrativo = administrativos.find_one({'_id': ObjectId(id)})
    administrativo['_id'] = str(administrativo['_id'])
    return render_template('edit_administrativo.html', administrativo=administrativo)

@app.route('/administrativos/<id>', methods=['DELETE'])
def delete_administrativo(id):
    administrativos.delete_one({'_id': ObjectId(id)})
    return jsonify({'msg': 'Administrativo eliminado'})

@app.route('/administrativos/json', methods=['GET'])
def get_all_administrativos_json():
    all_administrativos = list(administrativos.find({}))
    for administrativo in all_administrativos:
        administrativo['_id'] = str(administrativo['_id'])
    return jsonify(all_administrativos)

# Ruta para obtener todos los administrativos y renderizar el template
@app.route('/administrativos', methods=['GET'])
def get_all_administrativos():
    all_administrativos = list(administrativos.find({}).sort("nombre", 1))
    for administrativo in all_administrativos:
        administrativo['_id'] = str(administrativo['_id'])
    return render_template('administrativos.html', administrativos=all_administrativos)

# Generar Excel
@app.route('/export', methods=['POST'])
def export_data():
    data = request.json
    selected_columns = data['columns'].split(',')
    collection_name = data['collection']
    export_format = data.get('format', 'xlsx')
    collection = db[collection_name]

    cursor = collection.find({})
    df = pd.DataFrame(list(cursor))

    # Reemplazar valores NaN e Inf en el DataFrame antes de exportarlo
    df = df.replace([float('inf'), -float('inf')], 0).fillna('')

    # Verificar si hay datos en el DataFrame
    if df.empty:
        return {"message": "No hay datos para exportar"}, 400

    # Mapeos de nombres de columnas
    profesores_column_mapping = {
        "Nombre": "nombre",
        "Profesion": "profesion",
        "Adscripción": "adscripcion",
        "Fecha de ingreso": "fecha_ingreso",
        "Tiempo determinado": "tiempo_determinado",
        "Periodo actual": "periodo_actual",
        "Horas A": "horas_a",
        "Horas B": "horas_b",
        "Total de Horas": "total_horas",
    }

    column_mapping = profesores_column_mapping if collection_name == 'profesores' else {}

    # Filtrar las columnas seleccionadas y mapear a nombres internos
    selected_columns_db = [column_mapping[col] for col in selected_columns if col in column_mapping]

    # Verificar columnas dinámicas para asignación de horas
    for i in range(1, 9):
        for field in [f'asignatura{i}', f'grupo{i}', f'horas{i}']:
            if field not in df.columns:
                df[field] = None  # Agregar columna vacía si no existe

    output = BytesIO()

    if export_format == 'xlsx':
        # Exportar a formato XLSX
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('Sheet1')
            format_header = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#f4cccc'})
            format_centered = workbook.add_format({'align': 'center'})

            # Definir encabezados
            headers = [col for col in selected_columns if col in profesores_column_mapping]
            if 'asignacion_horas_frente_grupo' in selected_columns:
                headers.append('Asignación de Horas Frente a Grupo')
            worksheet.write_row('A1', headers, format_header)

            # Combinar celdas para 'Asignación de Horas Frente a Grupo' si está seleccionado
            if 'asignacion_horas_frente_grupo' in selected_columns:
                worksheet.merge_range('K1:S1', 'Asignación de Horas Frente a Grupo', format_header)

            # Escribir datos
            row_num = 1
            for index, row in df.iterrows():
                col_num = 0
                for col in headers:
                    if col in profesores_column_mapping:
                        db_col = profesores_column_mapping[col]
                        if db_col in row:
                            worksheet.write(row_num, col_num, row[db_col])
                    col_num += 1

                if 'asignacion_horas_frente_grupo' in selected_columns:
                    asignaciones_headers = ['Asignatura', 'Grupo', 'Horas', 'Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes', 'Sabado']
                    worksheet.write_row(row_num, col_num, asignaciones_headers, format_header)

                    for i in range(1, 9):
                        asignatura = row.get(f'asignatura{i}', '')
                        grupo = row.get(f'grupo{i}', '')
                        horas = row.get(f'horas{i}', 0) if row.get(f'horas{i}') else 0

                        if asignatura or grupo or horas:
                            worksheet.write_string(row_num + i, col_num, asignatura)
                            worksheet.write_string(row_num + i, col_num + 1, grupo)
                            worksheet.write_number(row_num + i, col_num + 2, float(horas))
                            worksheet.write_string(row_num + i, col_num + 3, f"{row.get(f'hora_inicio{i}1', '')} {row.get(f'hora_fin{i}1', '')}")
                            worksheet.write_string(row_num + i, col_num + 4, f"{row.get(f'hora_inicio{i}2', '')} {row.get(f'hora_fin{i}2', '')}")
                            worksheet.write_string(row_num + i, col_num + 5, f"{row.get(f'hora_inicio{i}3', '')} {row.get(f'hora_fin{i}3', '')}")
                            worksheet.write_string(row_num + i, col_num + 6, f"{row.get(f'hora_inicio{i}4', '')} {row.get(f'hora_fin{i}4', '')}")
                            worksheet.write_string(row_num + i, col_num + 7, f"{row.get(f'hora_inicio{i}5', '')} {row.get(f'hora_fin{i}5', '')}")
                            worksheet.write_string(row_num + i, col_num + 8, f"{row.get(f'hora_inicio{i}6', '')} {row.get(f'hora_fin{i}6', '')}")

                    worksheet.write_string(row_num + 9, col_num + 1, "Total:")
                    worksheet.write_number(row_num + 9, col_num + 2, float(row.get('total_horas_grupo', 0)) if row.get('total_horas_grupo') else 0)

                row_num += 10  # Avanzar 10 filas para el próximo profesor

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{collection_name}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    elif export_format == 'csv':
        # Exportar a formato CSV, incluyendo las asignaciones de horas
        csv_output = StringIO()
        if 'asignacion_horas_frente_grupo' in selected_columns:
            # Agregar columnas dinámicas para la asignación de horas
            for i in range(1, 9):
                df[f'Asignatura {i}'] = df[f'asignatura{i}']
                df[f'Grupo {i}'] = df[f'grupo{i}']
                df[f'Horas {i}'] = df[f'horas{i}']

        df.to_csv(csv_output, index=False, encoding='utf-8-sig')
        csv_output.seek(0)
        return send_file(
            csv_output,
            as_attachment=True,
            download_name=f"{collection_name}.csv",
            mimetype='text/csv',
        )

    return {"message": "Formato no soportado"}, 400




# COLUMNAS DE LAS TABLAS
@app.route('/columns/<collection_name>', methods=['GET'])
def get_columns(collection_name):
    if collection_name == 'profesores':
        columns = [
            "Nombre",
            "Profesion",
            "Adscripción",
            "Fecha de ingreso",
            "Tiempo determinado",
            "Periodo actual",
            "Horas A",
            "Horas B",
            "Total de Horas",
            "Cargo"
        ]
    elif collection_name == 'asignaturas':
        columns = [
            # Añade las columnas de la colección 'asignaturas' aquí
            "Nombre",  # Ejemplo
            "Horas",  # Ejemplo
            # Agrega más columnas según corresponda
        ]
    else:
        return jsonify({"error": "Colección no encontrada"}), 404

    return jsonify(columns)


@app.route('/index')
def index():
    return render_template('index.html')

# Ruta para la página principal
@app.route('/')
def principal():
    return render_template('principal.html')

# Rutas para reporteador
@app.route('/reporteador')
def reporteador():
    return render_template('exportacion/exportar.html')

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/upload-images", methods=["POST"])
def upload_images():
    if "header" in request.files:
        header = request.files["header"]
        if header and allowed_file(header.filename):
            header.save(os.path.join(app.config["UPLOAD_FOLDER"], "Encabezado1.PNG"))

    if "footer" in request.files:
        footer = request.files["footer"]
        if footer and allowed_file(footer.filename):
            footer.save(os.path.join(app.config["UPLOAD_FOLDER"], "PieDePagina1.PNG"))

    return jsonify({'msg': 'Imágenes actualizadas correctamente'}), 200

# Ruta para actualizar el texto en la celda A4
@app.route('/update-text', methods=['POST'])
def update_text():
    nuevo_texto = request.form.get('nuevo_texto', '')
    if not nuevo_texto:
        return jsonify({'msg': 'Error: No se proporcionó un texto válido'}), 400
    
    # Guardar el nuevo texto en un archivo de configuración
    with open("static/src/texto_a4.txt", "w", encoding="utf-8") as file:
        file.write(nuevo_texto)
    
    return jsonify({'msg': 'Texto actualizado correctamente'}), 200

# Ruta para exportar los profesores seleccionados usando la plantilla con 32 hojas EXCEL
@app.route('/export-selected', methods=['POST'])
def export_selected():
    profesor_ids = request.form.getlist('profesor_ids')
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [profesores.find_one({'_id': ObjectId(profesor_id)}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas
    template_path = "static/src/Plantilla_pie_reducido_2cm.xlsx"  # Cambia esta ruta si es necesario
    workbook = openpyxl.load_workbook(template_path)

    # Cargar imágenes
    header_image = Image("static/src/Encabezado1.PNG")
    footer_image = Image("static/src/PieDePagina1.PNG")

    # Establecer borde inferior en la celda A44
    border_bottom = Border(bottom=Side(style='thin'))

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 desde el archivo
    texto_a4_path = "static/src/texto_a4.txt"
    if os.path.exists(texto_a4_path):
        with open(texto_a4_path, "r", encoding="utf-8") as file:
            texto_a4 = file.read().strip()
    else:
        texto_a4 = ""

    # Obtener el nombre y cargo del encargado de dirección académica
    encargado = administrativos.find_one({
        'cargo': {'$in': ["ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA"]}
    })
    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""

    # Obtener el nombre y cargo del encargado de dirección general
    direccion = administrativos.find_one({
        'cargo': {'$in': ["DIRECTORA GENERAL", "DIRECTOR GENERAL"]}
    })
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos
    cargo_mapping = {
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "MECATRÓNICA": ["JEFA DE DIVISIÓN DE ING. MECATRÓNICA", "JEFE DE DIVISIÓN DE ING. MECATRÓNICA"],
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. SISTEMAS COMPUTACIONALES"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        # Insertar imágenes en encabezado y pie de página
        sheet.add_image(header_image, "A1")
        sheet.add_image(footer_image, "A54")

        # Actualizar el texto en la celda A4
        sheet["A4"] = texto_a4

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de direccion
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)


        # Llenar las celdas con "H.T." basado en los distintivos de grupo 81
        for j in range(1, 9):  # Iteramos sobre las asignaturas 1 a 8
            distintivo_key = f"distintivo{j}"  # Nombre del campo en la base de datos
            distintivo_valor = profesor.get(distintivo_key, "")

            if distintivo_valor:  # Si hay un valor en el distintivo
                fila = 14 + j  # Mapea asignatura1 a fila 15, asignatura2 a fila 16, etc.
                columnas = { "1": "E", "2": "F", "3": "G", "4": "H", "5": "I", "6": "J" }

                # Verificamos que el valor del distintivo esté en el rango esperado (1-6)
                columna = columnas.get(str(distintivo_valor))
                if columna:
                    celda_destino = f"{columna}{fila}"  # Ejemplo: H15

                    # Obtener el valor actual de la celda (si existe)
                    valor_existente = sheet[celda_destino].value

                    #Concatenar " H.T." al final
                    sheet[celda_destino].value = f"{valor_existente} H.T. "


        # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado = encargado["nombre"] if encargado else ""
            cargo_encargado = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado
            sheet["A46"] = nombre_encargado
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_bottom
            sheet["B44"].border = border_bottom

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # Guardar el archivo en memoria
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Devolver el archivo actualizado como respuesta
    return send_file(
        output,
        as_attachment=True,
        download_name="Reporte_Profesores_Dinamico.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Función para convertir Excel a PDF usando win32com
def excel_to_pdf(input_excel_path, output_pdf_path):
    pythoncom.CoInitialize()  # Inicializa el entorno COM
    try:
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Ejecutar Excel en segundo plano

        workbook = excel.Workbooks.Open(os.path.abspath(input_excel_path))
        
        # Iterar por todas las hojas y deshabilitar encabezados/pies de página
        for sheet in workbook.Sheets:
            sheet.PageSetup.CenterFooter = ""  # Elimina el pie de página central
            sheet.PageSetup.LeftFooter = ""    # Elimina el pie de página izquierdo
            sheet.PageSetup.RightFooter = ""   # Elimina el pie de página derecho
            sheet.PageSetup.CenterHeader = ""  # Elimina el encabezado central
            sheet.PageSetup.LeftHeader = ""    # Elimina el encabezado izquierdo
            sheet.PageSetup.RightHeader = ""   # Elimina el encabezado derecho
        
        # Exportar como PDF sin encabezado ni pie de página
        workbook.ExportAsFixedFormat(0, os.path.abspath(output_pdf_path))
        workbook.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()  # Liberar el entorno COM

# Ruta para exportar los profesores seleccionados en PDF
@app.route('/export-selected-pdf', methods=['POST'])
def export_selected_pdf():
    profesor_ids = request.form.getlist('profesor_ids')
    print("Profesores seleccionados:", profesor_ids)  # Verificar si se reciben los IDs correctamente

    if not profesor_ids:
        return "Error: No se recibieron IDs de profesores", 400
    
    fecha_aplicacion = request.form.get('fechaAplicacion', '')
    consecutivo = request.form.get('consecutivo', '')

    # Formatear la fecha de aplicación a DD/MM/YYYY
    if fecha_aplicacion:
        fecha_aplicacion = datetime.strptime(fecha_aplicacion, "%Y-%m-%d").strftime("%d/%m/%Y")

    selected_profesores = [profesores.find_one({'_id': ObjectId(profesor_id)}) for profesor_id in profesor_ids]

    # Ruta de la plantilla con 32 hojas
    template_path = "static/src/Plantilla_pie_reducido_2cm.xlsx"  # Cambia esta ruta si es necesario
    workbook = openpyxl.load_workbook(template_path)

    # Cargar imágenes
    header_image = Image("static/src/Encabezado1.PNG")
    footer_image = Image("static/src/PieDePagina1.PNG")

    # Establecer borde inferior en la celda A44
    border_bottom = Border(bottom=Side(style='thin'))

    # Estilo de relleno gris oscuro y texto blanco con fuente Helvetica 7 negrita
    dark_gray_fill = PatternFill(start_color="757171", end_color="757171", fill_type="solid")
    white_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFF")

    # Leer el texto de la celda A4 desde el archivo
    texto_a4_path = "static/src/texto_a4.txt"
    if os.path.exists(texto_a4_path):
        with open(texto_a4_path, "r", encoding="utf-8") as file:
            texto_a4 = file.read().strip()
    else:
        texto_a4 = ""

    # Obtener el nombre y cargo del encargado de dirección académica
    encargado = administrativos.find_one({
        'cargo': {'$in': ["ENCARGADA DEL DESPACHO DE DIRECCIÓN ACADÉMICA", "ENCARGADO DEL DESPACHO DE DIRECCIÓN ACADÉMICA"]}
    })
    nombre_encargado = encargado["nombre"] if encargado else ""
    cargo_encargado = encargado["cargo"] if encargado else ""

    # Obtener el nombre y cargo del encargado de dirección general
    direccion = administrativos.find_one({
        'cargo': {'$in': ["DIRECTORA GENERAL", "DIRECTOR GENERAL"]}
    })
    nombre_direccion = direccion["nombre"] if direccion else ""
    cargo_direccion = direccion["cargo"] if direccion else ""

    # Definir el mapeo de carreras a cargos administrativos
    cargo_mapping = {
        "INDUSTRIAL": ["JEFA DE DIVISIÓN DE ING. INDUSTRIAL", "JEFE DE DIVISIÓN DE ING. INDUSTRIAL"],
        "ELECTRÓNICA": ["JEFA DE DIVISIÓN DE ING. ELECTRÓNICA", "JEFE DE DIVISIÓN DE ING. ELECTRÓNICA"],
        "MECATRÓNICA": ["JEFA DE DIVISIÓN DE ING. MECATRÓNICA", "JEFE DE DIVISIÓN DE ING. MECATRÓNICA"],
        "SISTEMAS COMPUTACIONALES": ["JEFA DE DIVISIÓN DE ING. SISTEMAS COMPUTACIONALES", "JEFE DE DIVISIÓN DE ING. SISTEMAS COMPUTACIONALES"],
        "INFORMÁTICA": ["JEFA DE DIVISIÓN DE ING. INFORMÁTICA", "JEFE DE DIVISIÓN DE ING. INFORMÁTICA"],
        "ADMINISTRACIÓN": ["JEFA DE DIVISIÓN DE ING. ADMINISTRACIÓN", "JEFE DE DIVISIÓN DE ING. ADMINISTRACIÓN"]
    }

    # Llenar las hojas necesarias con los datos seleccionados
    for i, profesor in enumerate(selected_profesores):
        if i >= len(workbook.sheetnames):  # Evitar exceder el número de hojas disponibles
            break
        sheet = workbook[workbook.sheetnames[i]]

        #Insertar imagenes en encabezado y pie de pagina
        sheet.add_image(header_image, "A1")
        sheet.add_image(footer_image, "A54")

        # Actualizar el texto en la celda A4
        sheet["A4"] = texto_a4

        # Insertar nombre y cargo del encargado
        sheet["C45"] = nombre_encargado
        sheet["C46"] = cargo_encargado

        # Insertar nombre y cargo de direccion
        sheet["C52"] = nombre_direccion
        sheet["C53"] = cargo_direccion

        # Mapear datos en la hoja
        sheet["A7"] = profesor.get("nombre", "")
        sheet["E7"] = profesor.get("profesion", "")
        sheet["A9"] = profesor.get("adscripcion", "")
        sheet["E9"] = profesor.get("fecha_ingreso", "")
        sheet["H9"] = profesor.get("tiempo_indeterminado", "")
        sheet["A11"] = profesor.get("periodo_actual", "")
        sheet["F11"] = profesor.get("horas_a", 0)
        sheet["H11"] = profesor.get("horas_b", 0)
        sheet["J11"] = profesor.get("total_horas", 0)

        # Llenar las celdas de fecha de aplicación y consecutivo
        sheet["G40"] = fecha_aplicacion
        sheet["G41"] = consecutivo
        
        sheet["G44"] = profesor.get("nombre", "")

        # Total horas grupo
        sheet["D23"] = profesor.get("total_horas_grupo", 0)

        # Asignaturas y horarios
        for j in range(1, 9):
            row = 14 + j
            sheet[f"B{row}"] = profesor.get(f"asignatura{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupo{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horas{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicio{j}1', '')} {profesor.get(f'hora_fin{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicio{j}2', '')} {profesor.get(f'hora_fin{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicio{j}3', '')} {profesor.get(f'hora_fin{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicio{j}4', '')} {profesor.get(f'hora_fin{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicio{j}5', '')} {profesor.get(f'hora_fin{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicio{j}6', '')} {profesor.get(f'hora_fin{j}6', '')}"

        # Llenar las celdas A15-A22 basándose en C15-C22
        for row in range(15, 23):
            grupo = sheet[f"C{row}"].value
            if grupo:
                if str(grupo).startswith("1"):
                    sheet[f"A{row}"] = "INDUSTRIAL"
                elif str(grupo).startswith("4"):
                    sheet[f"A{row}"] = "SISTEMAS COMPUTACIONALES"

        # Asignaturas especiales y horarios
        for j in range(1, 9):
            row = 25 + j
            sheet[f"A{row}"] = profesor.get(f"carreraE{j}", "")
            sheet[f"B{row}"] = profesor.get(f"asignaturaE{j}", "")
            sheet[f"C{row}"] = profesor.get(f"grupoE{j}", "")
            sheet[f"D{row}"] = profesor.get(f"horasE{j}", 0)
            sheet[f"E{row}"] = f"{profesor.get(f'hora_inicioE{j}1', '')} {profesor.get(f'hora_finE{j}1', '')}"
            sheet[f"F{row}"] = f"{profesor.get(f'hora_inicioE{j}2', '')} {profesor.get(f'hora_finE{j}2', '')}"
            sheet[f"G{row}"] = f"{profesor.get(f'hora_inicioE{j}3', '')} {profesor.get(f'hora_finE{j}3', '')}"
            sheet[f"H{row}"] = f"{profesor.get(f'hora_inicioE{j}4', '')} {profesor.get(f'hora_finE{j}4', '')}"
            sheet[f"I{row}"] = f"{profesor.get(f'hora_inicioE{j}5', '')} {profesor.get(f'hora_finE{j}5', '')}"
            sheet[f"J{row}"] = f"{profesor.get(f'hora_inicioE{j}6', '')} {profesor.get(f'hora_finE{j}6', '')}"

        # Total horas grupo especial
        sheet["D34"] = profesor.get("total_horasE_grupo", 0)

        # Datos del cargo
        sheet["B37"] = profesor.get("cargo", "")
        sheet["C37"] = profesor.get("vigenciaCargo", "")
        sheet["D37"] = profesor.get("horasC", 0)

        # Horarios del cargo
        sheet["E37"] = f"{profesor.get('hora_inicioC11', '')} {profesor.get('hora_finC11', '')}"
        sheet["F37"] = f"{profesor.get('hora_inicioC12', '')} {profesor.get('hora_finC12', '')}"
        sheet["G37"] = f"{profesor.get('hora_inicioC13', '')} {profesor.get('hora_finC13', '')}"
        sheet["H37"] = f"{profesor.get('hora_inicioC14', '')} {profesor.get('hora_finC14', '')}"
        sheet["I37"] = f"{profesor.get('hora_inicioC15', '')} {profesor.get('hora_finC15', '')}"
        sheet["J37"] = f"{profesor.get('hora_inicioC16', '')} {profesor.get('hora_finC16', '')}"

        # Total horas generales
        sheet["D39"] = profesor.get("total_horas", 0)

        # Obtener horas B desde la celda H11
        horas_b = int(sheet["H11"].value) if sheet["H11"].value else 0

        # Listas de celdas a evaluar para cambiar el fondo
        horario_seccion_1 = [f"{col}{row}" for row in range(15, 23) for col in "EFGHIJ"]
        horario_seccion_2 = [f"{col}{row}" for row in range(26, 34) for col in "EFGHIJ"]

        # Contador de horas contabilizadas
        horas_restantes = horas_b

        # Función para procesar horarios y cambiar el fondo de las celdas
        def procesar_horarios(seccion):
            nonlocal horas_restantes
            for celda in seccion:
                if horas_restantes <= 0:
                    break
                valor = sheet[celda].value
                if valor and " " in valor:  # Verifica si hay datos y si tiene un horario válido
                    try:
                        # Extraer solo las partes de la hora ignorando cualquier "H.T."
                        partes = valor.split(" ")
                        if len(partes) >= 2:  # Asegurar que tenemos al menos dos partes (inicio y fin)
                            inicio = partes[0]  # Ejemplo: "07:00"
                            fin = partes[1]  # Ejemplo: "09:00"
                            
                            # Extraer la hora en formato numérico
                            horas = int(fin[:2]) - int(inicio[:2])  # Obtener la cantidad de horas

                            if horas <= horas_restantes:
                                sheet[celda].fill = dark_gray_fill  # Aplicar sombreado gris oscuro
                                sheet[celda].font = white_font  # Cambiar color de fuente a blanco
                                horas_restantes -= horas  # Restar horas utilizadas
                    except ValueError:
                        continue  # Evita errores si el formato no es el esperado

        # Aplicar la lógica a ambas secciones
        procesar_horarios(horario_seccion_1)
        if horas_restantes > 0:
            procesar_horarios(horario_seccion_2)


        # Llenar las celdas con "H.T." basado en los distintivos de grupo 81
        for j in range(1, 9):  # Iteramos sobre las asignaturas 1 a 8
            distintivo_key = f"distintivo{j}"  # Nombre del campo en la base de datos
            distintivo_valor = profesor.get(distintivo_key, "")

            if distintivo_valor:  # Si hay un valor en el distintivo
                fila = 14 + j  # Mapea asignatura1 a fila 15, asignatura2 a fila 16, etc.
                columnas = { "1": "E", "2": "F", "3": "G", "4": "H", "5": "I", "6": "J" }

                # Verificamos que el valor del distintivo esté en el rango esperado (1-6)
                columna = columnas.get(str(distintivo_valor))
                if columna:
                    celda_destino = f"{columna}{fila}"  # Ejemplo: H15

                    # Obtener el valor actual de la celda (si existe)
                    valor_existente = sheet[celda_destino].value

                    #Concatenar " H.T." al final
                    sheet[celda_destino].value = f"{valor_existente} H.T. "

         # Obtener los valores de A15-A22 y A26-A33 sin valores vacíos
        carreras_detectadas = sorted({sheet[f"A{row}"].value for row in range(15, 23) if sheet[f"A{row}"].value})
        carreras_detectadas += sorted({sheet[f"A{row}"].value for row in range(26, 34) if sheet[f"A{row}"].value})
        carreras_detectadas = list(set(carreras_detectadas))  # Eliminar duplicados

        # Si hay un solo valor en carreras_detectadas
        if len(carreras_detectadas) == 1 and carreras_detectadas[0] in cargo_mapping:
            carrera = carreras_detectadas[0]
            encargado = administrativos.find_one({
                'cargo': {'$in': cargo_mapping[carrera]}
            })
            nombre_encargado = encargado["nombre"] if encargado else ""
            cargo_encargado = encargado["cargo"] if encargado else ""
            
            # Asignar valores en las celdas
            sheet["A43"] = cargo_encargado
            sheet["A46"] = nombre_encargado
        
        # Si hay exactamente dos valores en carreras_detectadas
        elif len(carreras_detectadas) == 2:
            carrera1, carrera2 = carreras_detectadas
            encargado1 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera1, [])}})
            encargado2 = administrativos.find_one({'cargo': {'$in': cargo_mapping.get(carrera2, [])}})

            nombre1 = encargado1["nombre"] if encargado1 else ""
            cargo1 = encargado1["cargo"] if encargado1 else ""
            nombre2 = encargado2["nombre"] if encargado2 else ""
            cargo2 = encargado2["cargo"] if encargado2 else ""

            # Asignar valores en las celdas
            sheet["A43"] = cargo1
            sheet["A44"] = nombre1
            sheet["A45"] = cargo2
            sheet["A46"] = nombre2
            sheet["A44"].border = border_bottom
            sheet["B44"].border = border_bottom

    # Eliminar las hojas no utilizadas
    for i in range(len(selected_profesores), len(workbook.sheetnames)):
        del workbook[workbook.sheetnames[-1]]

    # Guardar el archivo Excel temporalmente para la conversión a PDF
    temp_excel_path = "temp_reporte.xlsx"
    workbook.save(temp_excel_path)
    workbook.close()

    # Convertir el archivo Excel a PDF
    pdf_path = "Reporte_Profesores.pdf"
    excel_to_pdf(temp_excel_path, pdf_path)

    # Enviar el archivo PDF como respuesta
    return send_file(
        pdf_path,
        as_attachment=True,
        download_name="Reporte_Profesores.pdf",
        mimetype="application/pdf"
    )

if __name__ == '__main__':
    app.run(debug=True)
